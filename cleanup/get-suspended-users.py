from github import Github
import openpyxl
import datetime
import os

# Authentication is defined via github.Auth
from github import Auth

github_token = os.environ['GITHUB_TOKEN']
# using an access token
auth = Auth.Token(github_token)

filename = "tmp/user_status.xlsx"
age_threashold = 180
try:
    os.remove(filename)
except OSError:
    pass

# First create a Github instance:

# Public Web Github
g = Github(auth=auth)

# Github Enterprise with custom hostname
g = Github(base_url="https://github.ibm.com/api/v3", auth=auth)

workbook = openpyxl.Workbook()
with open("cleanup/orgs.txt", "r") as f:
    org_names = f.readlines()

worksheet = workbook.create_sheet("inactive_users")
row_num = 1
# Add columns for team name, username, email, and repo name
worksheet.cell(row=1, column=1).value = "Org"
worksheet.cell(row=1, column=2).value = "user"
worksheet.cell(row=1, column=3).value = "Suspended Date"


for org_name in org_names:
    org_name = org_name.strip()
    org = g.get_organization(org_name)
    print (f'Org: {org_name}')
    for member in org.get_members():
        if str(member.suspended_at) != 'None':
            age = datetime.date.today() - member.suspended_at.date()
            # print (f'{member.name} - {member.suspended_at} - {age}')
            if (age.days > age_threashold):
                print (f'Marking {member.name} for removal from {org_name}')
                row_num += 1
                worksheet.cell(row=row_num, column=1).value = org_name
                worksheet.cell(row=row_num, column=2).value = member.name
                worksheet.cell(row=row_num, column=2).hyperlink = member.html_url
                worksheet.cell(row=row_num, column=2).style = "Hyperlink"
                worksheet.cell(row=row_num, column=3).value = str(member.suspended_at)
        else:
            print ('skipping...')

workbook.save(filename)

# To close connections after use
g.close()